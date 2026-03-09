"""
Import バーガンディ在庫.xlsx into wc_beverages via API
- Converts half-width katakana to full-width
- Maps wine type + region → category_id
- All items → store_id = 'burgundy'
"""
import json, unicodedata, sys, os
from openpyxl import load_workbook
try:
    import requests
except ImportError:
    os.system(f'{sys.executable} -m pip install requests')
    import requests

EXCEL = r'C:\Users\RyojiHayashi\Downloads\バーガンディ在庫.xlsx'
API = 'https://wine-compass.vercel.app/api/import'
STORE_ID = 'burgundy'

# --- Category mapping ---
# Red (parent=3) sub-categories by region
RED_REGION_MAP = {
    'ブルゴーニュ': 15,  # Burgundy Red
    'ボルドー': 16,      # Bordeaux Red
    'ローヌ': 17,        # Rhône Red
    'イタリア': 19,      # Italy Red
}
RED_FRANCE_OTHER = 18    # Other France Red
RED_OTHER = 20           # Other Red

# White (parent=2) sub-categories by region
WHITE_REGION_MAP = {
    'ブルゴーニュ': 8,   # Burgundy White
    'ボルドー': 9,       # Bordeaux White
    'ロワール': 10,      # Loire White
    'アルザス': 11,      # Alsace White
    'イタリア': 13,      # Italy White
}
WHITE_FRANCE_OTHER = 12  # Other France White
WHITE_OTHER = 14         # Other White

# French country names (normalized)
FRENCH_COUNTRIES = {'フランス'}
ITALY_COUNTRIES = {'イタリア'}

def hw_to_fw(text):
    """Half-width katakana → full-width + general NFKC normalization"""
    if not text:
        return text
    return unicodedata.normalize('NFKC', str(text)).strip()

def get_category_id(wine_type, region, country):
    """Map wine type + region/country → category_id"""
    t = hw_to_fw(wine_type) if wine_type else ''
    r = hw_to_fw(region) if region else ''
    c = hw_to_fw(country) if country else ''
    is_france = c in FRENCH_COUNTRIES
    is_italy = c in ITALY_COUNTRIES or 'イタリア' in r

    if t == '赤':
        for key, cat_id in RED_REGION_MAP.items():
            if key in r:
                return cat_id
        if is_italy:
            return 19
        if is_france:
            return RED_FRANCE_OTHER
        return RED_OTHER

    if t == '白':
        for key, cat_id in WHITE_REGION_MAP.items():
            if key in r:
                return cat_id
        if is_italy:
            return 13
        if is_france:
            return WHITE_FRANCE_OTHER
        return WHITE_OTHER

    if t in ('泡白', '泡ロゼ'):
        return 1  # Champagne / Sparkling

    if t == 'ロゼ':
        return 4

    if t == '蒸留酒':
        return 6  # Beer & Spirits

    if t == 'オレンジ':
        return WHITE_OTHER  # Orange wine → Other White

    if t == '甘口':
        return WHITE_OTHER  # Sweet → Other White

    if t == '黄色':
        return WHITE_OTHER  # Vin Jaune → Other White

    return RED_OTHER  # Default fallback

def parse_vintage(v):
    if v is None or v == '' or v == 0:
        return None
    s = str(v).strip()
    if s.upper() == 'NV':
        return None
    try:
        n = int(float(s))
        if 1900 <= n <= 2100:
            return n
    except (ValueError, TypeError):
        pass
    return None

def parse_int(v):
    if v is None or v == '':
        return None
    try:
        n = int(float(str(v)))
        return n if n > 0 else None
    except (ValueError, TypeError):
        return None

def parse_qty(v):
    if v is None or v == '':
        return 0
    try:
        return max(0, int(float(str(v))))
    except (ValueError, TypeError):
        return 0

def parse_size(v):
    if v is None or v == '':
        return 750
    try:
        n = int(float(str(v)))
        if n <= 0:
            return 750
        if n > 30000:
            return 750  # Cap unreasonable sizes (SMALLINT limit + real-world max)
        return n
    except (ValueError, TypeError):
        return 750


# --- Read Excel ---
print(f'Loading {EXCEL}...')
wb = load_workbook(EXCEL, read_only=True, data_only=True)
ws = wb[wb.sheetnames[0]]

rows = list(ws.iter_rows(min_row=1, values_only=True))
wb.close()
print(f'Read {len(rows)} rows (including header)')

# Header row
header = rows[0]
# Columns: 倉庫名(0), 国名(1), 産地名(2), 商品コード(3), Ｖt(4), 商品名1(5), 商品名2(6),
#           タイプ(7), 容量(8), 上代(9), 仕入額(10), 在庫本数(11), インポーター(12), [total_cost(13)]

items = []
skipped = 0

for row in rows[1:]:
    if not row or len(row) < 12:
        skipped += 1
        continue

    warehouse = row[0]
    country = row[1]
    region = row[2]
    code = row[3]
    vintage_raw = row[4]
    name1 = row[5]   # Wine name (half-width katakana)
    name2 = row[6]   # Producer (half-width katakana)
    wine_type = row[7]
    volume = row[8]
    retail = row[9]
    cost = row[10]
    qty = row[11]
    importer = row[12] if len(row) > 12 else None

    # Skip empty rows
    if not name1:
        skipped += 1
        continue

    # Convert half-width → full-width
    name_fw = hw_to_fw(name1)
    producer_fw = hw_to_fw(name2) if name2 else None
    region_fw = hw_to_fw(region) if region else None
    country_fw = hw_to_fw(country) if country else None
    warehouse_fw = hw_to_fw(warehouse) if warehouse else None
    importer_fw = hw_to_fw(importer) if importer else None
    type_fw = hw_to_fw(wine_type) if wine_type else None

    cat_id = get_category_id(wine_type, region, country)

    # Build notes: warehouse + importer + code
    notes_parts = []
    if warehouse_fw:
        notes_parts.append(f'倉庫: {warehouse_fw}')
    if importer_fw:
        notes_parts.append(f'仕入: {importer_fw}')
    if code:
        notes_parts.append(f'Code: {code}')

    item = {
        'name': name_fw,
        'name_kana': name_fw,  # Already katakana
        'producer': producer_fw,
        'vintage': parse_vintage(vintage_raw),
        'region': f'{country_fw} / {region_fw}' if country_fw and region_fw else (region_fw or country_fw or None),
        'appellation': region_fw,
        'size_ml': parse_size(volume),
        'quantity': parse_qty(qty),
        'price': parse_int(retail),
        'cost_price': parse_int(cost),
        'category_id': cat_id,
        'notes': ' / '.join(notes_parts) if notes_parts else None,
    }
    items.append(item)

print(f'Parsed {len(items)} items ({skipped} skipped)')

# --- Category stats ---
cat_counts = {}
for it in items:
    cid = it['category_id']
    cat_counts[cid] = cat_counts.get(cid, 0) + 1
print(f'Category distribution:')
CAT_NAMES = {
    1:'Champagne', 4:'Rose', 6:'Spirits',
    8:'Burg White', 9:'Bordeaux White', 10:'Loire White', 11:'Alsace White',
    12:'Other FR White', 13:'Italy White', 14:'Other White',
    15:'Burg Red', 16:'Bordeaux Red', 17:'Rhone Red',
    18:'Other FR Red', 19:'Italy Red', 20:'Other Red',
}
for cid, cnt in sorted(cat_counts.items()):
    print(f'  {cid:>2} ({CAT_NAMES.get(cid, "?")}): {cnt}')

total_bottles = sum(it['quantity'] for it in items)
print(f'Total bottles: {total_bottles}')

# --- Sample items ---
print(f'\nSample items:')
for it in items[:5]:
    print(f'  {it["vintage"] or "NV"} {it["name"]} / {it["producer"]} qty={it["quantity"]} cat={it["category_id"]}')

# --- Send to API ---
print(f'\nSending {len(items)} items to {API} (store={STORE_ID})...')
resp = requests.post(API, json={'store_id': STORE_ID, 'items': items}, timeout=120)
print(f'Status: {resp.status_code}')
result = resp.json()
print(json.dumps(result, indent=2, ensure_ascii=False))
